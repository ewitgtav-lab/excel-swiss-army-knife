from __future__ import annotations

import re
from io import BytesIO
from typing import Literal

import numpy as np
import pandas as pd
import streamlit as st


st.set_page_config(page_title="Ultimate Excel Automator", layout="wide")


# ---------------------------
# Caching + File Readers
# ---------------------------


def _freeze_uploads(files) -> list[tuple[str, bytes]]:
    # Convert UploadedFile objects to stable, hashable inputs for caching.
    return [(f.name, f.getvalue()) for f in files]


def _sanitize_df(df: pd.DataFrame) -> pd.DataFrame:
    # Ensure consistent/serializable headers immediately (Streamlit Arrow / PyArrow stability).
    df = df.copy()
    df.columns = df.columns.map(lambda c: str(c).strip())
    return df


def _arrow_safe_for_display(df: pd.DataFrame) -> pd.DataFrame:
    """
    Streamlit's dataframe uses Arrow. Mixed-type 'object' columns (e.g. int + str, bytes + int)
    can crash Arrow conversion. For display only, coerce object columns to pandas 'string'.
    """
    if df.empty:
        return df

    safe = df.copy()
    for col in safe.columns:
        if safe[col].dtype == "object":
            safe[col] = safe[col].astype("string")
    return safe


@st.cache_data(show_spinner=False)
def _read_csv_bytes(name: str, data: bytes) -> pd.DataFrame:
    bio = BytesIO(data)
    # Let pandas infer delimiter; Excel-heavy users often have standard CSV.
    df = pd.read_csv(bio)
    return _sanitize_df(df)


@st.cache_data(show_spinner=False)
def _read_excel_bytes(name: str, data: bytes) -> pd.DataFrame:
    bio = BytesIO(data)
    # Excel users often have data on non-first sheets; read all sheets and combine.
    sheets = pd.read_excel(bio, sheet_name=None)
    dfs = []
    for sheet_name, df in (sheets or {}).items():
        if isinstance(df, pd.DataFrame) and not df.empty:
            df = df.copy()
            df.insert(0, "_sheet", str(sheet_name))
            dfs.append(df)
    if not dfs:
        # Fallback: some Excel files confuse pandas (blank header rows, formatting-only sheets, etc).
        # Use openpyxl to extract the "used range" values.
        try:
            import openpyxl

            wb = openpyxl.load_workbook(BytesIO(data), data_only=True, read_only=True)
            extracted: list[pd.DataFrame] = []
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                max_row = ws.max_row or 0
                max_col = ws.max_column or 0
                if max_row <= 1 or max_col <= 0:
                    continue

                values = [
                    [ws.cell(row=r, column=c).value for c in range(1, max_col + 1)]
                    for r in range(1, max_row + 1)
                ]
                # Drop trailing completely-empty rows
                while values and all(v is None or (isinstance(v, str) and not v.strip()) for v in values[-1]):
                    values.pop()
                if len(values) <= 1:
                    continue

                header = [str(v).strip() if v is not None else "" for v in values[0]]
                body = values[1:]
                df2 = pd.DataFrame(body, columns=header)
                df2.insert(0, "_sheet", str(sheet_name))
                # If header is empty (common), fall back to generic column names
                if all(h == "" for h in header):
                    df2.columns = ["_sheet"] + [f"col_{i+1}" for i in range(df2.shape[1] - 1)]
                extracted.append(df2)

            if extracted:
                combined = pd.concat(extracted, ignore_index=True)
                return _sanitize_df(combined)
        except Exception:
            pass

        return pd.DataFrame()
    combined = pd.concat(dfs, ignore_index=True)
    return _sanitize_df(combined)


@st.cache_data(show_spinner=False)
def _read_docx_bytes(name: str, data: bytes) -> pd.DataFrame:
    from docx import Document

    bio = BytesIO(data)
    doc = Document(bio)

    # Flatten all tables to a single dataframe (best-effort).
    rows: list[list[str]] = []
    max_cols = 0
    for t in doc.tables:
        for r in t.rows:
            row = [c.text for c in r.cells]
            max_cols = max(max_cols, len(row))
            rows.append(row)

    if not rows:
        return pd.DataFrame()

    # Pad jagged rows to rectangular.
    padded = [r + [""] * (max_cols - len(r)) for r in rows]
    df = pd.DataFrame(padded)
    return _sanitize_df(df)


@st.cache_data(show_spinner=False)
def _read_pdf_bytes(name: str, data: bytes) -> pd.DataFrame:
    import pdfplumber

    bio = BytesIO(data)
    dfs: list[pd.DataFrame] = []
    with pdfplumber.open(bio) as pdf:
        for page in pdf.pages:
            table = page.extract_table()
            if table:
                dfs.append(pd.DataFrame(table))

    if not dfs:
        return pd.DataFrame()

    df = pd.concat(dfs, ignore_index=True)
    return _sanitize_df(df)


@st.cache_data(show_spinner="Sharpening the Knife...", ttl=3600)
def load_and_combine(frozen_uploads: list[tuple[str, bytes]]) -> pd.DataFrame:
    dfs: list[pd.DataFrame] = []

    for name, data in frozen_uploads:
        lower = name.lower()
        try:
            if lower.endswith(".csv"):
                dfs.append(_read_csv_bytes(name, data))
            elif lower.endswith(".pdf"):
                dfs.append(_read_pdf_bytes(name, data))
            elif lower.endswith(".docx"):
                dfs.append(_read_docx_bytes(name, data))
            else:
                dfs.append(_read_excel_bytes(name, data))
        except Exception as e:
            st.error(f"Error reading {name}: {e}")

    dfs = [d for d in dfs if isinstance(d, pd.DataFrame) and not d.empty]
    if not dfs:
        return pd.DataFrame()

    df = pd.concat(dfs, ignore_index=True)
    return _sanitize_df(df)


@st.cache_data(show_spinner="Sharpening the Knife...", ttl=3600)
def load_and_combine_with_report(
    frozen_uploads: list[tuple[str, bytes]],
) -> tuple[pd.DataFrame, list[dict]]:
    """
    Same as load_and_combine, but returns an extraction report for debugging
    (useful for mobile upload edge cases).
    """
    report: list[dict] = []
    dfs: list[pd.DataFrame] = []

    for name, data in frozen_uploads:
        lower = name.lower()
        item: dict = {
            "name": name,
            "type": lower.rsplit(".", 1)[-1] if "." in lower else "",
            "bytes": len(data) if data is not None else 0,
            "rows": 0,
            "cols": 0,
            "error": None,
            "note": None,
        }
        try:
            if item["bytes"] < 10_000 and lower.endswith((".xlsx", ".xlsm", ".xls")):
                item["note"] = "File is very small for Excel; may be empty or upload truncated."
            if lower.endswith(".csv"):
                d = _read_csv_bytes(name, data)
            elif lower.endswith(".pdf"):
                d = _read_pdf_bytes(name, data)
            elif lower.endswith(".docx"):
                d = _read_docx_bytes(name, data)
            else:
                d = _read_excel_bytes(name, data)

            if isinstance(d, pd.DataFrame):
                item["rows"] = int(d.shape[0])
                item["cols"] = int(d.shape[1])
                if not d.empty:
                    dfs.append(d)
        except Exception as e:
            item["error"] = str(e)

        report.append(item)

    if not dfs:
        return pd.DataFrame(), report
    df = pd.concat(dfs, ignore_index=True)
    return _sanitize_df(df), report


# ---------------------------
# Export
# ---------------------------


@st.cache_data(show_spinner="Building Excel download...", ttl=3600)
def _df_to_xlsx_bytes(df: pd.DataFrame) -> bytes:
    output = BytesIO()
    with pd.ExcelWriter(output, engine="xlsxwriter") as writer:
        df.to_excel(writer, index=False)
    return output.getvalue()


# ---------------------------
# Vectorized utilities
# ---------------------------


def _to_clean_numeric(series: pd.Series) -> pd.Series:
    # Removes common currency/format chars and converts to numeric.
    s = series.astype(str).str.replace(r"[$,\s]", "", regex=True)
    return pd.to_numeric(s, errors="coerce")


def _format_scientific_notation(series: pd.Series) -> pd.Series:
    """
    Goal: convert float-y numeric values into human Excel-friendly strings without scientific notation.
    Vectorized strategy:
    - parse numeric
    - if value is "close to integer" -> Int64 -> string
    - else keep a trimmed decimal representation (fixed precision then strip)
    """
    numeric = pd.to_numeric(series, errors="coerce")
    out = series.astype(object).copy()

    mask = numeric.notna()
    if not mask.any():
        return out

    n = numeric[mask].astype(float)
    rounded = np.round(n)
    is_intlike = np.isclose(n, rounded, rtol=0.0, atol=1e-9)

    # Int-like: render as integer string.
    int_vals = pd.Series(rounded.astype("int64"), index=numeric[mask].index)
    out.loc[int_vals.index[is_intlike]] = int_vals.loc[int_vals.index[is_intlike]].astype(str)

    # Non-int: fixed precision, then strip trailing zeros/dot (still vectorized).
    non_int_idx = int_vals.index[~is_intlike]
    if len(non_int_idx) > 0:
        s = pd.Series(n[~is_intlike], index=non_int_idx)
        rendered = s.round(10).astype(str).str.replace(r"\.?0+$", "", regex=True)
        out.loc[non_int_idx] = rendered

    return out


# ---------------------------
# State + Reset
# ---------------------------


def _ensure_state():
    st.session_state.setdefault("main_df", None)
    st.session_state.setdefault("xlsx_bytes", None)


def _hard_reset():
    st.cache_data.clear()
    st.session_state.clear()
    st.rerun()


def _set_df(df: pd.DataFrame):
    st.session_state.main_df = df
    st.session_state.xlsx_bytes = None  # invalidate export cache in state


# ---------------------------
# App
# ---------------------------


_ensure_state()

with st.sidebar:
    st.header("🛠️ Tool Belt")
    if st.button("🧨 Hard Reset (clear memory + cache)", width="stretch"):
        _hard_reset()

    if st.button("♻️ Reset Data Only", width="stretch"):
        st.session_state.main_df = None
        st.session_state.xlsx_bytes = None
        st.rerun()

    st.divider()
    st.header("📣 Support")
    st.link_button("🪲 Report a Bug", "https://forms.gle/rVE2KkorZX4iqWNq7")
    st.link_button("☕ Support my Work", "https://paypal.me/GewishCatedrilla")


st.title("🛠️ The Excel Swiss Army Knife")


if st.session_state.main_df is None:
    files = st.file_uploader(
        "Upload Files",
        type=["xlsx", "csv", "pdf", "docx"],
        accept_multiple_files=True,
    )
    if files:
        # On some mobile browsers, an immediate rerun after upload can lose the uploader state.
        # So we load + set session_state in the same run.
        frozen = _freeze_uploads(files)
        loaded, report = load_and_combine_with_report(frozen)
        if loaded is None or loaded.empty:
            st.error(
                "Upload received, but no tables/rows could be extracted. "
                "Try a different file, or export your Excel to CSV and re-upload."
            )
            with st.expander("Debug details (extraction report)"):
                st.write(report)
        else:
            st.session_state.main_df = loaded
            # proceed without rerun


df: pd.DataFrame | None = st.session_state.main_df
if df is None or df.empty:
    st.info("Upload one or more files to begin.")
    st.stop()


st.write(f"### 🔍 Preview ({len(df)} rows)")
st.dataframe(_arrow_safe_for_display(df.head(10)), width="stretch")

tabs = st.tabs(
    [
        "🧮 Aggregator",
        "🎯 Mapper",
        "🧹 Cleaner",
        "🕵️ Detective",
        "⏰ Time Math",
        "🔄 Shifter",
        "✅ Validator",
        "📂 Word",
    ]
)


# TAB 1: AGGREGATOR
with tabs[0]:
    st.header("Smart Aggregator")
    with st.form("agg_form"):
        g_col = st.selectbox("Group by:", df.columns, key="agg_g")
        s_col = st.selectbox("Sum Column:", df.columns, key="agg_s")
        go = st.form_submit_button("Generate Summary")
    if go:
        try:
            s = _to_clean_numeric(df[s_col])
            value_name = s_col if g_col != s_col else f"{s_col}_sum"
            summary = (
                df.assign(_s=s)
                .groupby(g_col, sort=False)["_s"]
                .sum(min_count=1)
                .reset_index(name=value_name)
            )
            st.dataframe(summary, width="stretch")
        except Exception as e:
            st.error(f"Aggregator failed: {e}")


# TAB 2: MAPPER
with tabs[1]:
    st.header("Categorization Logic")
    with st.form("map_form"):
        m_col = st.selectbox("Scan Column:", df.columns, key="map_c")
        keyword = st.text_input("If text contains:")
        target = st.text_input("Then assign this category:", value="Uncategorized")
        go = st.form_submit_button("Apply Mapping")
    if go:
        try:
            if not keyword:
                st.warning("Enter a keyword to match.")
            else:
                working = df.copy()
                if "Category" not in working.columns:
                    working["Category"] = "Uncategorized"
                mask = working[m_col].astype(str).str.contains(re.escape(keyword), case=False, na=False)
                working.loc[mask, "Category"] = target or "Uncategorized"
                _set_df(working)
                st.success("Logic Applied! Check preview above.")
                st.rerun()
        except Exception as e:
            st.error(f"Mapper failed: {e}")


# TAB 3: CLEANER
with tabs[2]:
    st.header("Formatting Fixer")
    with st.form("clean_form"):
        c_col = st.selectbox("Target Column:", df.columns, key="clean_c")
        c_opt: Literal[
            "Scientific Notation", "Remove Symbols", "Proper Case", "Trim Whitespace"
        ] = st.radio(
            "Fix Type:",
            ["Scientific Notation", "Remove Symbols", "Proper Case", "Trim Whitespace"],
            horizontal=True,
        )
        go = st.form_submit_button("Run Fixer")
    if go:
        try:
            working = df.copy()
            if c_opt == "Remove Symbols":
                working[c_col] = working[c_col].astype(str).str.replace(r"[$\-,%]", "", regex=True)
            elif c_opt == "Proper Case":
                working[c_col] = working[c_col].astype(str).str.title()
            elif c_opt == "Trim Whitespace":
                working[c_col] = working[c_col].astype(str).str.strip()
            else:
                working[c_col] = _format_scientific_notation(working[c_col])

            _set_df(working)
            st.success("Cleaned!")
            st.rerun()
        except Exception as e:
            st.error(f"Cleaner failed: {e}")


# TAB 4: DETECTIVE
with tabs[3]:
    st.header("Duplicate Detective")
    with st.form("dupe_form"):
        d_cols = st.multiselect("Match rows on these columns:", df.columns, default=[])
        go = st.form_submit_button("Identify Duplicates")
    if go:
        try:
            if not d_cols:
                st.warning("Pick at least one column to match on.")
            else:
                dupes = df[df.duplicated(subset=d_cols, keep=False)]
                if not dupes.empty:
                    st.warning(f"Found {len(dupes)} duplicates.")
                    st.dataframe(dupes.astype(str), width="stretch")
                else:
                    st.success("No duplicates found!")
        except Exception as e:
            st.error(f"Detective failed: {e}")


# TAB 5: TIME MATH
with tabs[4]:
    st.header("Time & Labor Math")
    with st.form("time_form"):
        h_col = st.selectbox("Select Hours Column:", df.columns, key="tm_h")
        m_col = st.selectbox("Select Minutes Column:", df.columns, key="tm_m")
        out_col = st.text_input("Output column name:", value="Total_Hours_Decimal")
        go = st.form_submit_button("Combine to Decimal Hours")
    if go:
        try:
            working = df.copy()
            h = pd.to_numeric(working[h_col], errors="coerce").fillna(0)
            m = pd.to_numeric(working[m_col], errors="coerce").fillna(0)
            working[out_col] = h + (m / 60.0)
            _set_df(working)
            st.success(f"Created '{out_col}' column!")
            st.rerun()
        except Exception as e:
            st.error(f"Time Math failed: {e}")


# TAB 6: SHIFTER
with tabs[5]:
    st.header("Format Shifter")
    with st.form("shift_form"):
        s_opt = st.selectbox("Convert to:", ["Word-Ready CSV", "HTML Report", "JSON"])
        go = st.form_submit_button("Prepare Conversion")
    if go:
        try:
            if s_opt == "Word-Ready CSV":
                st.download_button(
                    "📥 Download CSV",
                    df.to_csv(index=False).encode("utf-8"),
                    "for_word.csv",
                    width="stretch",
                )
            elif s_opt == "HTML Report":
                st.download_button(
                    "📥 Download HTML",
                    df.to_html(index=False).encode("utf-8"),
                    "report.html",
                    width="stretch",
                )
            else:
                st.download_button(
                    "📥 Download JSON",
                    df.to_json(orient="records").encode("utf-8"),
                    "export.json",
                    width="stretch",
                )
        except Exception as e:
            st.error(f"Shifter failed: {e}")


# TAB 7: VALIDATOR
with tabs[6]:
    st.header("Rule Validator (Crash-Proof)")
    with st.form("val_form"):
        v_col = st.selectbox("Column to Validate:", df.columns, key="val_c")
        v_type = st.radio(
            "Rule:",
            ["Must be Numeric", "Must be Email", "Cannot be Empty"],
            horizontal=True,
        )
        go = st.form_submit_button("Validate Now")
    if go:
        try:
            if v_type == "Must be Numeric":
                clean = pd.to_numeric(df[v_col], errors="coerce")
                errors = df[clean.isna() & df[v_col].astype(str).str.strip().ne("")]
            elif v_type == "Must be Email":
                email_ok = df[v_col].astype(str).str.match(r"^[^@\s]+@[^@\s]+\.[^@\s]+$", na=False)
                errors = df[~email_ok]
            else:
                errors = df[df[v_col].astype(str).str.strip() == ""]

            if not errors.empty:
                st.error(f"🚨 Found {len(errors)} invalid rows!")
                st.dataframe(errors.astype(str), width="stretch")
            else:
                st.success("✅ All data in this column is valid!")
        except Exception as e:
            st.error(f"Validator failed: {e}")


# TAB 8: WORD
with tabs[7]:
    st.header("Word Table Import")
    st.info("If you upload `.docx`, tables are extracted and included in the combined dataset.")


st.divider()

col1, col2 = st.columns([1, 2], vertical_alignment="center")
with col1:
    if st.button("⚡ Prepare Excel Download", width="stretch"):
        try:
            st.session_state.xlsx_bytes = _df_to_xlsx_bytes(df)
            st.success("Excel download prepared.")
        except Exception as e:
            st.session_state.xlsx_bytes = None
            st.error(f"Excel export failed: {e}")

with col2:
    xbytes = st.session_state.get("xlsx_bytes")
    st.download_button(
        "📥 DOWNLOAD COMPLETED SWISS ARMY FILE",
        xbytes if xbytes else b"",
        "fixed_data.xlsx",
        disabled=not bool(xbytes),
        width="stretch",
    )

